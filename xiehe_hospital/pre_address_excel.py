#coding=utf8

from openpyxl import load_workbook
from datetime import datetime
import json


class Xiehe(object):

    SPLIT = '-'

    PATIENT_ID_COL = 'C'

    _col_and_name = {
        'B':u'姓名',
        'C':u'病案号',
        'D':u'备注',
        'E':u'临床结局',
        'F':u'主管医生',
        'G':u'是否治疗',
        'H':u'定位日期',
        'I':u'副反应'+SPLIT+u'1周',
        'J':u'副反应'+SPLIT+u'2周',
        'K':u'副反应'+SPLIT+u'3周',
        'L':u'副反应'+SPLIT+u'4周',
        'M':u'副反应'+SPLIT+u'5周',
        'N':u'副反应'+SPLIT+u'6周',
        'O':u'副反应'+SPLIT+u'7周',
        'P':u'腹泻',
        'Q':u'直肠炎',
        'R':u'尿频',
        'S':u'恶心',
        'T':u'呕吐',
        'U':u'腹痛',
        'V':u'乏力',
        'W':u'疗后一周妇查',
        'X':u'疗后二周妇查',
        'Y':u'疗后三周妇查',
        'Z':u'血液毒性'+SPLIT+u'WBC白细胞',
        'AA':u'血液毒性'+SPLIT+u'NEUT中性粒细胞',
        'AB':u'血液毒性'+SPLIT+u'HGB血红蛋白',
        'AC':u'血液毒性'+SPLIT+u'PLT血小板',
        'AD':u'血液毒性'+SPLIT+u'淋巴细胞',
        'AE':u'血液毒性'+SPLIT+u'WBC分级',
        'AF':u'血液毒性'+SPLIT+u'NEUT分级',
        'AG':u'血液毒性'+SPLIT+u'HGB分级',
        'AH':u'血液毒性'+SPLIT+u'PLT分级',
        'AI':u'血液毒性'+SPLIT+u'LY分级',
        'AJ':u'血液毒性'+SPLIT+u'备注',
        'AK':u'年龄',
        'AL':u'地区',
        'AM':u'电话1',
        'AN':u'电话2',
        'AO':u'电话3',
        'AP':u'电话4',
        'AQ':u'既往史',
        'AR':u'病理'+SPLIT+u'分级',
        'AS':u'病理'+SPLIT+u'分型',
        'AT':u'病理号',
        'AU':u'不详时间',
        'AV':u'妇科检查'+SPLIT+u'外阴',
        'AW':u'妇科检查'+SPLIT+u'阴道',
        'AX':u'妇科检查'+SPLIT+u'宫颈占位大小',
        'AY':u'妇科检查'+SPLIT+u'性质',
        'AZ':u'妇科检查'+SPLIT+u'子宫',
        'BA':u'妇科检查'+SPLIT+u'宫旁',
        'BB':u'妇科检查'+SPLIT+u'三合诊',
        'BC':u'分期',
        'BD':u'血常规'+SPLIT+u'WBC白细胞',
        'BE':u'血常规'+SPLIT+u'NEUT中性粒细胞',
        'BF':u'血常规'+SPLIT+u'LY淋巴细胞',
        'BG':u'血常规'+SPLIT+u'HGB血红蛋白',
        'BH':u'血常规'+SPLIT+u'PLT血小板',
        'BI':u'肾肝功'+SPLIT+u'ALT',
        'BJ':u'肾肝功'+SPLIT+u'Cr',
        'BK':u'SCC',
        'BL':u'CA125',
        'BM':u'感染'+SPLIT+u'HBsAg',
        'BN':u'感染'+SPLIT+u'HBsAb',
        'BO':u'感染'+SPLIT+u'HBeAg',
        'BP':u'感染'+SPLIT+u'HbeAb',
        'BQ':u'感染'+SPLIT+u'HBcAb',
        'BR':u'感染'+SPLIT+u'HCV',
        'BS':u'感染'+SPLIT+u'TP',
        'BT':u'感染'+SPLIT+u'HIV',
        'BU':u'HPV',
        'BV':u'肾血流功能显像'+SPLIT+u'印象',
        'BW':u'肾血流功能显像'+SPLIT+u'GFR',
        'BX':u'肾血流功能显像'+SPLIT+u'右肾GFR',
        'BY':u'肾血流功能显像'+SPLIT+u'左肾GFR',
        'BZ':u'MRI报告'+SPLIT+u'宫颈病变大小cm',
        'CA':u'MRI报告'+SPLIT+u'描述',
        'CB':u'MRI报告'+SPLIT+u'外阴',
        'CC':u'MRI报告'+SPLIT+u'阴道',
        'CD':u'MRI报告'+SPLIT+u'子宫',
        'CE':u'MRI报告'+SPLIT+u'宫旁',
        'CF':u'MRI报告'+SPLIT+u'盆腔积液',
        'CG':u'MRI报告'+SPLIT+u'淋巴结区域',
        'CH':u'MRI报告'+SPLIT+u'淋巴结大小',
        'CI':u'MRI报告'+SPLIT+u'其他',
        'CJ':u'MRI片',
        'CK':u'原发部位',
        'CL':u'CT'+SPLIT+u'胸',
        'CM':u'CT'+SPLIT+u'腹部',
        'CN':u'CT'+SPLIT+u'腹主动脉淋巴结',
        'CO':u'CT'+SPLIT+u'盆腔淋巴位置',
        'CP':u'不详',
        'CQ':u'CT片',
        'CR':u'PET/CT'+SPLIT+u'原发情况',
        'CS':u'PET/CT'+SPLIT+u'淋巴结转移情况',
        'CT':u'PET/CT'+SPLIT+u'远处转移情况',
        'CU':u'PET/CT'+SPLIT+u'其他',
        'CV':u'影像'+SPLIT+u'胸',
        'CW':u'影像'+SPLIT+u'腹',
        'CX':u'影像'+SPLIT+u'盆',
        'CY':u'骨扫描',
        'CZ':u'腹主动脉旁淋巴结转移'+SPLIT+u'有/无',
        'DA':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'有/无',
        'DB':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'数目',
        'DC':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'单侧/双侧',
        'DD':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'最大直径',
        'DE':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'左侧最大',
        'DF':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'右侧最大',
        'DG':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'髂总'+SPLIT+u'有/无',
        'DH':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'髂总'+SPLIT+u'单侧/双侧',
        'DI':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'髂总'+SPLIT+u'最大直径',
        'DJ':u'盆腔淋巴结转移(CT严格短径1cm)'+SPLIT+u'髂总'+SPLIT+u'腹膜后照射的上界',
        'DK':u'先期化疗'+SPLIT+u'方案及计量',
        'DL':u'先期化疗'+SPLIT+u'开始时间',
        'DM':u'先期化疗'+SPLIT+u'结束时间',
        'DN':u'先期化疗'+SPLIT+u'疗程数',
        'DO':u'先期化疗'+SPLIT+u'方式(静脉/动脉)',
        'DP':u'不详',
        'DQ':u'放疗方案'+SPLIT+u'开始时间',
        'DR':u'放疗方案'+SPLIT+u'结束时间',
        'DS':u'放疗方案'+SPLIT+u'持续时间',
        'DT':u'放疗方案'+SPLIT+u'放疗方式',
        'DU':u'靶区范围'+SPLIT+u'描述',
        'DV':u'靶区范围'+SPLIT+u'盆腔',
        'DW':u'靶区范围'+SPLIT+u'腹主动脉旁',
        'DX':u'靶区范围'+SPLIT+u'腹股沟',
        'DY':u'靶区范围'+SPLIT+u'靶区剂量',
        'DZ':u'靶区范围'+SPLIT+u'淋巴结加量',
        'EA':u'靶区范围'+SPLIT+u'剂量',
        'EB':u'靶区范围'+SPLIT+u'有无宫旁补量',
        'EC':u'靶区范围'+SPLIT+u'单侧/双侧',
        'ED':u'靶区范围'+SPLIT+u'宫旁补量',
        'EE':u'靶区范围'+SPLIT+u'内照射剂量及次数',
        'EF':u'靶区范围'+SPLIT+u'备注',
        'EG':u'化疗'+SPLIT+u'方案',
        'EH':u'化疗'+SPLIT+u'次数',
        'EI':u'化疗'+SPLIT+u'开始时间',
        'EJ':u'化疗'+SPLIT+u'结束时间',
        'EK':u'化疗'+SPLIT+u'备注',
        'EL':u'不详',
        'EM':u'第一次复查'+SPLIT+u'时间',
        'EN':u'第一次复查'+SPLIT+u'距放疗结束',
        'EO':u'第一次复查'+SPLIT+u'SCC',
        'EP':u'第一次复查'+SPLIT+u'CA125',
        'EQ':u'第一次复查'+SPLIT+u'疗效',
        'ER':u'第一次复查'+SPLIT+u'副反应',
        'ES':u'第一次复查'+SPLIT+u'追加放疗',
        'ET':u'第一次复查'+SPLIT+u'复发部位',
        'EU':u'第一次复查'+SPLIT+u'转移部位',
        'EV':u'第一次复查'+SPLIT+u'治疗建议',
        'EW':u'第一次复查'+SPLIT+u'治疗',
        'EX':u'第一次复查'+SPLIT+u'再次复查',
        'EY':u'第一次复查'+SPLIT+u'备注',
        'IR':u'不详',
        'IS':u'结局'+SPLIT+u'不治疗',
        'IT':u'结局'+SPLIT+u'复发转移',
        'IU':u'结局'+SPLIT+u'不明确',
        'IV':u'结局'+SPLIT+u'有肿瘤残留',
        'IW':u'结局'+SPLIT+u'死亡',
        'IX':u'结局'+SPLIT+u'其他原因死亡',
        'IY':u'粗分期',
        'IZ':u'有无腹主动脉淋巴结转移',
        'JA':u'有无盆腔淋巴结转移',
        'JB':u'既往史'+SPLIT+u'高血压',
        'JC':u'既往史'+SPLIT+u'糖尿病',
        'JD':u'既往史'+SPLIT+u'脑血管病史(脑梗/出血等)',
        'JE':u'既往史'+SPLIT+u'其他妇科疾病史(卵巢癌/子宫肌瘤等)',
        'JF':u'既往史'+SPLIT+u'各种手术史',
        'JG':u'既往史'+SPLIT+u'药物过敏史'
    }

    _sheet = {
        'sheet_name':'Sheet1',
        'begin_row':4,
        'end_row':666
    }

    _replace_punctuation = [
        ('，', ','),
        ('（', '('),
        ('）', ')'),
        ('。', '.'),
        ('；', ';'),
        ('：', ':'),
        ('“', '"'),
        ("”", '"')
    ]

    def __init__(self, file_path):
        self.file_path = file_path
        self.processed_content = self.transform_spreadsheet_to_json()

    def transform_spreadsheet_to_json(self):
        """
        Transform data in spreadsheet to json format
        :return: dict, json format representation of spreadsheet data
        """
        wb = load_workbook(filename=self.file_path, data_only=True)
        ws = wb.get_sheet_by_name(self._sheet['sheet_name'])
        ret = {}
        for row in range(self._sheet['begin_row'], self._sheet['end_row']+1):
            patient_id = ws[self.PATIENT_ID_COL+str(row)].value
            ret[patient_id] = { t[0]:t[1] for t in self.process_each_row(ws, row)}
        return ret

    def process_each_row(self, sheet, row):
        """
        Each row data indicates a series of info of a patient
        :param sheet: spreadsheet to work on
        :param row: row index number
        :return: tuple, ( a, b )
                   a: str, column index
                   b: str, processed content of the corresponding element in spreadsheet
        """
        for col in self._col_and_name.iterkeys():
            val = sheet[col+str(row)].value
            if isinstance(val, datetime):
                yield (col, val.strftime('%Y/%m/%d'))
            elif isinstance(val, unicode):
                yield (col, self.clean_description_feature(val.encode('utf-8')))
            else:
                yield (col, val)

    def clean_description_feature(self, val):
        """
        clean punctuation or other dirties in each patients description feature
        :param val: description content
        :return: str, cleaned description content
        """
        for t in self._replace_punctuation:
            val = val.replace(t[0], t[1])
        return val

    def filter_patient_sample(self):
        """
        filter patient sample
        :return: dict, filtered patient samples
        """
        samples = self.processed_content
        for k,v in samples.items():
            if samples[k]['IS'] == '是':
                del samples[k]
        self.processed_content = samples

    def transform_AV_BB(self):
        samples = self.processed_content
        for k,v in samples.iteritems():
            samples[k]['AV_binary'] = ''
            if samples[k]['AV']:
                samples[k]['AV_binary'] = 0 if samples[k]['AV']=='(-)' else 1
        self.processed_content = samples

    def transform_BK(self):
        samples = self.processed_content
        for k,v in samples.iteritems():
            if '70' in str(samples[k]['BK']):
                samples[k]['BK'] = '70'
        self.processed_content = samples

    def transform_DK(self):
        samples = self.processed_content
        for k,v in samples.iteritems():
            samples[k]['DK'] = '1' if samples[k]['DK'] else '0'
        self.processed_content = samples

    def transform_DS(self):
        samples = self.processed_content
        for k,v in samples.iteritems():
            if 'VALUE' in str(samples[k]['DS']) or '持续时间' in str(samples[k]['DS']):
                samples[k]['DS'] = None
        self.processed_content = samples

    def transform_JB_JG(self):
        samples = self.processed_content
        for k,v in samples.iteritems():
            for col in ['JB','JC','JD','JE','JF','JG']:
                samples[k][col] = 1 if samples[k][col]=='有' else 0
        self.processed_content = samples




if __name__ == '__main__':
    from ipdb import set_trace as st
    # st(context=21)
    xiehe = Xiehe('./data/support/协和放疗科交付数据-v3.xlsx')
    xiehe.filter_patient_sample()
    xiehe.transform_AV_BB()
    xiehe.transform_BK()
    xiehe.transform_DK()
    xiehe.transform_DS()
    xiehe.transform_JB_JG()
    ret = xiehe.processed_content
    json.dump(ret, open('./data/output/output.dat','w'), indent=4, sort_keys=True, ensure_ascii=False)
